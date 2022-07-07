from email import header
from sqlalchemy import create_engine    
import pandas as pd
from openpyxl import load_workbook

def connect_sql_server():
    info_siem_name = ['rain2002kr','Showme0022^^','siemens2020.synology.me',5307 ,'PLAN_DB' ]

    USER = info_siem_name[0]
    PASSWORD= info_siem_name[1]
    HOST= info_siem_name[2]
    PORT =info_siem_name[3]
    DATABASE= info_siem_name[4]

    connection_string = "mysql+pymysql://%s:%s@%s:%s/%s" % (USER, PASSWORD, HOST, PORT, DATABASE)
    _engine = create_engine(connection_string)
    
    return _engine

def make_df_from_excel(sht_name):
    src =  r'D:\97. 업무공유파일\000. 계획\01. 2022 일일계획표.xlsx'
    try:
        df1 = pd.read_excel(src, sheet_name= sht_name)
        cur_date = df1.columns[1]
        return df1, cur_date
    except :
        ERROR = f'ERROR READ {sht_name} NAME from EXCEL FILE'
        print(f"CODE : {ERROR} ") 

def make_df_plan_sum(engine, org_df, cur_date):
    d_cols_jobs_total = []
    d_rows_jobs_total = [0]
    v_start_jobs_total = 11
    v_end_jobs_total = 18 
    v_step = 2
    tb_name ='tb_plan_sum'
    
    for i in range( v_start_jobs_total, v_end_jobs_total + 1) :
        d_cols_jobs_total.append(i)
    try:
        df_plan_sum = org_df.iloc[21:22,d_cols_jobs_total].dropna()
        df_plan_sum['날짜'] = cur_date
        # df_plan_sum.to_sql(tb_name, con=engine, if_exists= 'append', index=False)
        return df_plan_sum
    
    except :
        ERROR = f'make_df_plan_sum : ERROR MAKE DateFrame from the Sequence'
        print(f"CODE : {ERROR} ") 

def make_df_from_sql(engine,tb_name, cur_date):
    # query = f'SELECT * FROM {tb_name} WHERE 날짜="{cur_date}";'
    query = f'SELECT * FROM {tb_name};'
    try:
        df_sql_sum = pd.read_sql(query, con=engine)
        con1 = df_sql_sum['날짜'].duplicated(keep='last')
        df1_sql_sum = df_sql_sum[~con1]
        df1_sql_sum.to_sql(tb_name, con=engine,if_exists='replace', index=False)

        return df1_sql_sum
    except :
        ERROR = f'ERROR READ {tb_name} NAME from MariaDB Platform'
        print(f"CODE : {ERROR} ") 
        print(f"CODE : {query} ") 

def make_df_plan_jobs(engine, org_df, cur_date):

    d_cols_jobs_total = []
    d_rows_jobs_total = [0]
    v_start = 11
    v_end = 24 
    v_step = 2
    JOBS = []
    jobs = []
    tb_name ='tb_job'
        
    try:
        
        for i in range(v_start, v_end, v_step) :
            j = i + 1
            JOBS.append([i,j])
        
        job_name = ["업무", "계획", "개발", "필수", "펀", "교육", "이동"]
        start_row = 24

        for i in range(0,len(job_name)):
            df_jobs = org_df.iloc[start_row:,JOBS[i]].dropna()
            df_jobs = df_jobs.reset_index(drop=True)
            df_jobs_n = df_jobs.T.reset_index(drop=True)
            df_jobs_n.reset_index()
            jobs.append(df_jobs_n)    
        
        for i in range(0,len(jobs)):
            jobs[i].columns = jobs[i].iloc[0]
            jobs[i] = jobs[i].drop([jobs[i].index[0]])
            # jobs[i]['날짜'] = org_df.columns[1]
            # jobs[i].to_sql(tb_name+ str(i), con=engine, if_exists= 'append', index=False)
        
        return jobs
    
    except :
        ERROR = f'make_df_plan_jobs : ERROR MAKE DateFrame from the Sequence'
        print(f"CODE : {ERROR} ") 

def read_dfs_from_sql(engine, tb_name , cur_date):
    tb_name = 'tb_job'
    src_path = r"D:\97. 업무공유파일\000. 계획/"
    tar_path = src_path
    file =  '03. 시간분석테이블.xlsx'
    
    sht_name = tb_name
    df_jobs = []
    for i in range(0, 7):
        df_sql_jobs = make_df_from_sql(engine, tb_name + str(i), cur_date)
        
        df = df_sql_jobs
        save_excel(src_path, tar_path , file , df, sht_name)

        df_jobs.append(df_sql_jobs)
    return df_jobs

def save_excel(df, sht_name):
    src_path = r"D:\97. 업무공유파일\000. 계획/"
    tar_path = src_path
    file =  '03. 시간분석테이블.xlsx'

    book = load_workbook(src_path + file)
    writer = pd.ExcelWriter(tar_path + file, engine= 'openpyxl')
    writer.book = book
    # df.to_excel(writer, sheet_name= sht_name,startrow=writer.sheets[sht_name].max_row  , index=False, header=False)
    df.to_excel(writer, sheet_name= sht_name, index=False)
    writer.save()
    writer.close()

# mode='a'


