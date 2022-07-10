from dataclasses import replace
from sqlalchemy import create_engine    
import pandas as pd
from openpyxl import load_workbook
import common_pc


# 엑셀 저장 리팩터링 필요 
def main(start_date, cnt):
    engine = common_pc.connect_sql_server()

    if cnt == 0:
        date_b = '22'+str(start_date)
        sht_name = date_b 
        org_df, cur_date = common_pc.make_df_from_excel(sht_name)
        print(org_df)
        
        
        df_plan_sum = common_pc.make_df_plan_sum(org_df, cur_date)
        print(df_plan_sum)

        jobs = common_pc.make_df_plan_jobs(org_df, cur_date)
        print(jobs)

        tdf = pd.concat([jobs[1], jobs[2],jobs[3],jobs[4],jobs[5],jobs[6]],axis=1)
        jobs[0].insert(0, '날짜', org_df.columns[1])
        tdf.insert(0, '날짜', org_df.columns[1])

        print(tdf)
        print(jobs[0])
        common_pc.save_excel( tdf, sht_name = 'plan_details')
        common_pc.save_excel( jobs[0], sht_name = 'plan_works')

        ndf = common_pc.read_df_from_timetable('plan_works')
        print(ndf)
        ndf = common_pc.read_df_from_timetable('plan_details')
        print(ndf)


# def load_timesheet():
#     pd.read_excel()

def load_cur_excel(date):
    sht_name = str(date) 
    org_df, cur_date = common_pc.make_df_from_excel(sht_name)
    # print(org_df)
    
    df_plan_sum = common_pc.make_df_plan_sum(org_df, cur_date)
    # print(df_plan_sum)

    job0, tdf = common_pc.make_df_plan_jobs(org_df, cur_date)
    # print(job0)
    # print(tdf)
    return df_plan_sum, job0, tdf
    
def load_time_excel():
    src = r"D:\97. 업무공유파일\000. 계획\01. 시간분석테이블/03. 시간분석테이블.xlsx"
    df1 = pd.read_excel(src, sheet_name=0)
    df2 = pd.read_excel(src, sheet_name=1)
    df3 = pd.read_excel(src, sheet_name=2)
    return df1,df2,df3

def save_time_excel(df1,df2,df3):
    src = r"D:\97. 업무공유파일\000. 계획\01. 시간분석테이블/"
    file = "03. 시간분석테이블.xlsx"
    file_b = "03. 시간분석테이블-backup.xlsx"
    src_file = r"D:\97. 업무공유파일\000. 계획\01. 시간분석테이블/03. 시간분석테이블.xlsx"
    sht = []
    sht = read_sheets(src,file)
    print(sht)
    with pd.ExcelWriter(src_file) as writer:  
        df1.to_excel(writer, sheet_name=sht[0], index=False)
        df2.to_excel(writer, sheet_name=sht[1], index=False)
        df3.to_excel(writer, sheet_name=sht[2], index=False)

    with pd.ExcelWriter(src + file_b) as writer:  
        df1.to_excel(writer, sheet_name=sht[0], index=False)
        df2.to_excel(writer, sheet_name=sht[1], index=False)
        df3.to_excel(writer, sheet_name=sht[2], index=False)


# FUNCTION CODE : delete_target_sheet_by_number
# src_path = 이 안에 들어 있는 모든 엑셀 파일의 타겟 sheet를 삭제한다.
# excel        : path +엑셀파일.xlsx 
# tar_sht_number : 시트번호, 첫번째 0, 두번째 1 ... n번째 n
# comment      : 독립적으로 동작하고, 엑셀파일과 시트번호을 넣으면, 해당하는 시트가 삭제된다.      
###############################################################################################    
def delete_target_sheet_by_number(excel, tar_sht_number):
    sht = []
    wb = load_workbook(excel)

    for item in wb.sheetnames:
        sht.append(item)

    wb.remove(wb[sht[tar_sht_number]])
    wb.save(excel)

# FUNCTION CODE = read_sheets() 
def read_sheets(path, excel):
        i = -1
        wb = load_workbook(path + excel)
        shts = []
        for item in wb.sheetnames:
            i = i + 1
            shts.append(item)
        return shts

# start_date = '0702'
# cnt = 0
# main(start_date, cnt)


