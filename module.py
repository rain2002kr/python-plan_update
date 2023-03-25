# from pymongo import ASCENDING
from sqlalchemy import create_engine

# from sre_constants import SUCCESS
import pandas as pd
from openpyxl import load_workbook
import os

DEBUG_ON = 1
DEBUG_OFF = 0

debug = DEBUG_OFF
D_read_sheets = DEBUG_OFF
D_create_new_excel = DEBUG_OFF
D_load_time_excel = DEBUG_ON
D_load_cur_plan_excel = DEBUG_OFF
D_make_df_plan_sum = DEBUG_OFF
D_make_df_plan_jobs = DEBUG_OFF
D_make_df_from_arr = DEBUG_OFF
D_update_df_from_arr = DEBUG_OFF
D_save_time_excel = DEBUG_OFF
D_connect_sql_server = DEBUG_OFF
D_save_time_excel_to_server = DEBUG_OFF
D_load_time_excel_from_server = DEBUG_OFF
D_make_df_plan = DEBUG_OFF

# FUNCTION CODE = load_time_excel_from_server()
#
#
# return       : engine
# comment      : 서버 연결, ORM 이용해서 SQL 서버에 연결한다.
###############################################################################################
def load_time_excel_from_server():
    engine = connect_sql_server()
    # v_df_arr = load_time_excel()
    tb_name1 = "plan_sum"
    tb_name2 = "plan_job"
    tb_name3 = "plan_other"
    query1 = f"SELECT * FROM {tb_name1};"
    query2 = f"SELECT * FROM {tb_name2};"
    query3 = f"SELECT * FROM {tb_name3};"

    try:
        df_plan_sum = pd.read_sql(query1, con=engine)
        df_plan_job = pd.read_sql(query2, con=engine)
        df_plan_other = pd.read_sql(query3, con=engine)

        if debug or D_load_time_excel_from_server:
            SUCCESS = f"load_time_excel_from_server : SUCCESS"
            print(f"CODE : {SUCCESS} ")
            print(f"CODE : {df_plan_sum} ")
            print(f"CODE : {df_plan_job} ")
            print(f"CODE : {df_plan_other} ")

        return df_plan_sum, df_plan_job, df_plan_other

    except:
        if debug or D_load_time_excel_from_server:
            ERROR = f"load_time_excel_from_server : 서버 연결에 실패했습니다."
            print(f"CODE : {ERROR} ")


# FUNCTION CODE = connect_sql_server()
#
#
# return       : engine
# comment      : 서버 연결, ORM 이용해서 SQL 서버에 연결한다.
###############################################################################################
def connect_sql_server():
    USER = os.getenv("USER")
    PASSWORD = os.getenv("PASSWORD")
    HOST = os.getenv("HOST")
    PORT = os.getenv("PORT")
    DATABASE = os.getenv("DATABASE")

    try:
        connection_string = "mysql+pymysql://%s:%s@%s:%s/%s" % (
            USER,
            PASSWORD,
            HOST,
            PORT,
            DATABASE,
        )
        _engine = create_engine(connection_string)

        if debug or D_connect_sql_server:
            SUCCESS = f"connect_sql_server : SUCCESS"
            print(f"CODE : {SUCCESS} ")

        return _engine

    except:
        if debug or D_connect_sql_server:
            ERROR = f"connect_sql_server : 서버 연결에 실패했습니다."
            print(f"CODE : {ERROR} ")


# FUNCTION CODE = save_time_excel_to_server()
#
#
# return       :
# comment      :
###############################################################################################
def save_time_excel_to_server():
    engine = connect_sql_server()
    v_df_arr = load_time_excel()
    tb_name1 = "plan_sum"
    tb_name2 = "plan_job"
    tb_name3 = "plan_other"
    # query = f'SELECT * FROM {tb_name1};'
    try:

        if debug or D_save_time_excel_to_server:
            SUCCESS = f"save_time_excel_to_server : SUCCESS"
            print(f"CODE : {SUCCESS} ")
            print(f"CODE : {v_df_arr} ")

        v_df_arr[0].to_sql(tb_name1, con=engine, if_exists="replace", index=False)
        v_df_arr[1].to_sql(tb_name2, con=engine, if_exists="replace", index=False)
        v_df_arr[2].to_sql(tb_name3, con=engine, if_exists="replace", index=False)

    except:
        if debug or D_save_time_excel_to_server:
            ERROR = f"save_time_excel_to_server : ERROR "
            print(f"CODE : {ERROR} ")


# FUNCTION CODE = read_sheets()
# path         : 파일이 있는 경로를 지정한다.
# excel        : path +엑셀파일.xlsx
# return        : sheet 명을 배열 형태로 돌려준다.
# comment      : 독립적으로 동작하고, 경로와, 엑셀파일을 넣으면 해당하는 모든 시트를 읽는다.
###############################################################################################
def read_sheets(path, excel):
    i = -1
    wb = load_workbook(path + excel)
    shts = []
    try:
        for item in wb.sheetnames:
            i = i + 1
            shts.append(item)
        if debug or D_read_sheets:
            SUCCESS = f"read_sheets : SUCCESS"
            print(f"CODE : {SUCCESS} ")
            print(f"CODE : {shts} ")
        return shts
    except:
        if debug or D_read_sheets:
            ERROR = f"read_sheets : ERROR 파일확장자가 지정되지않음.EX) .xlsx"
            print(f"CODE : {ERROR} ")


# FUNCTION CODE : create_new_excel
# path = 이 안에 들어 있는 모든 엑셀 파일의 타겟 sheet를 삭제한다.
# file        : path +엑셀파일.xlsx
# tar_sht_number : 시트번호, 첫번째 0, 두번째 1 ... n번째 n
# comment      : 독립적으로 동작하고, 엑셀파일과 시트번호을 넣으면, 해당하는 시트가 삭제된다.
###############################################################################################
def create_new_excel():
    path = r"D:\97. 업무공유파일\000. 계획\01. 시간분석테이블/"
    path = r"W:\01. 업무\000. 계획\01. 시간분석테이블/"

    file = "03. 시간분석테이블.xlsx"
    df = pd.DataFrame()
    sht = ["통합", "업무", "기타"]
    try:
        with pd.ExcelWriter(path + file) as writer:
            df.to_excel(writer, sheet_name=sht[0], index=False)
            df.to_excel(writer, sheet_name=sht[1], index=False)
            df.to_excel(writer, sheet_name=sht[2], index=False)
        if debug or D_create_new_excel:
            SUCCESS = f"create_new_excel : SUCCESS"
            print(f"CODE : {SUCCESS} ")
            read_sheets(path, file)

    except:
        if debug or D_create_new_excel:
            ERROR = f"create_new_excel : ERROR 파일확장자가 지정되지않음.EX) .xlsx"
            print(f"CODE : {ERROR} ")


# FUNCTION CODE : load_time_excel
# path = 파일이 있는 경로명
# file = 파일명.xlsx
# src  = path + file
#
# comment : 시간분석테이블에 내용을 읽어온다.
###############################################################################################
def load_time_excel():
    path = r"D:\97. 업무공유파일\000. 계획\01. 시간분석테이블/"
    path = r"W:\01. 업무\000. 계획\01. 시간분석테이블/"
    file = "03. 시간분석테이블.xlsx"
    src = path + file
    try:
        df1 = pd.read_excel(src, sheet_name=0)
        df2 = pd.read_excel(src, sheet_name=1)
        df3 = pd.read_excel(src, sheet_name=2)

        if debug or D_load_time_excel:
            SUCCESS = f"load_time_excel : SUCCESS"
            print(f"CODE : {SUCCESS} ")
            print(f"CODE : {df1} ")
            print(f"CODE : {df2} ")
            print(f"CODE : {df3} ")

        return df1, df2, df3

    except:
        if debug or D_load_time_excel:
            ERROR = f"load_time_excel : ERROR 없는 시트에 접근 하였음"
            print(f"CODE : {ERROR} ")


# FUNCTION CODE : load_cur_plan_excel
# path = 파일이 있는 경로명
# file = 파일명.xlsx
# src  = path + file
# cnt_sht_num : 현재날짜 부터 몇개의 시트를 읽어올것인지 파라미터
# comment : 일일계획표에 지정된 시트를 모두 읽어와서 배열로 리턴한다.
###############################################################################################
def load_cur_plan_excel(cmd, sht_num_cnt=1, date="230101"):
    path = r"W:\01. 업무\000. 계획/"
    file = "01. 2023 일일계획표.xlsx"
    src = path + file
    frame = []
    try:
        v_str_command = cmd
        v_int_cnt_sht_num = int(sht_num_cnt)
        v_str_date = date
        v_int_start_sht_num = 4
        v_int_tar_sht_num = v_int_start_sht_num + v_int_cnt_sht_num

        if v_str_command == "load_shts_by_number":
            for i in range(v_int_start_sht_num, v_int_tar_sht_num):
                df = pd.read_excel(src, sheet_name=i)
                frame.append(df)
        elif v_str_command == "load_sht_by_date":
            df = pd.read_excel(src, sheet_name=v_str_date)
            frame.append(df)

        if debug or D_load_cur_plan_excel:
            SUCCESS = f"load_cur_plan_excel : SUCCESS"
            print(f"CODE : {SUCCESS} ")
            # 디버그시 현재 날짜 출력
            for i in range(0, len(frame)):
                print(f"\tREAD SHT DATE : {frame[i].columns[1]}")

        return frame

    except:
        if debug or D_load_cur_plan_excel:
            ERROR = f"load_cur_plan_excel : ERROR 없는 시트에 접근 하였음"
            print(f"CODE : {ERROR} ")


# FUNCTION CODE : make_df_plan_sum
# path
# file        :
# comment :
###############################################################################################
def make_df_plan_sum(org_df):
    SUM_JOBS_FST_COL = 11
    SUM_JOBS_END_COL = 18
    SUM_JOBS_FST_ROW = 21
    SUM_JOBS_END_ROW = 22
    v_arr_sum_jobs_col_addr = []
    v_str_cur_date = org_df.columns[1]

    try:
        for i in range(SUM_JOBS_FST_COL, SUM_JOBS_END_COL + 1):
            v_arr_sum_jobs_col_addr.append(i)

        df_plan_sum = org_df.iloc[
            SUM_JOBS_FST_ROW:SUM_JOBS_END_ROW, v_arr_sum_jobs_col_addr
        ].dropna()
        df_plan_sum.insert(0, "날짜", v_str_cur_date)
        df_plan_sum["날짜"] = df_plan_sum["날짜"].apply(
            lambda x: pd.Timestamp(x).strftime("%Y-%m-%d")
        )

        if debug or D_make_df_plan_sum:
            SUCCESS = f"make_df_plan_sum : SUCCESS"
            print(f"CODE : {SUCCESS} ")
            print(f"CODE : {df_plan_sum} ")

        return df_plan_sum

    except:
        if debug or D_make_df_plan_sum:
            ERROR = f"make_df_plan_sum : ERROR DateFrame from the Sequence"
            print(f"CODE : {ERROR} ")


def make_df_plan_jobs(org_df):
    v_arr_jobs_col_addr = []
    v_str_cur_date = org_df.columns[1]
    JOBS_FST_COL = 11
    JOBS_END_COL = 24
    JOBS_STEP = 2
    JOBS_FST_ROW = 24
    JOBS_QTY = 7
    v_arr_jobs = []

    try:
        # Get jobs columns array
        for i in range(JOBS_FST_COL, JOBS_END_COL, JOBS_STEP):
            j = i + 1
            v_arr_jobs_col_addr.append([i, j])

        # Get jobs dataframe array
        for i in range(0, JOBS_QTY):
            df_jobs = org_df.iloc[JOBS_FST_ROW:, v_arr_jobs_col_addr[i]].dropna()
            df_jobs = df_jobs.reset_index(drop=True).T.reset_index(drop=True)

            v_arr_jobs.append(df_jobs)

        # Set jobs column name with first row value
        for i in range(0, len(v_arr_jobs)):
            v_arr_jobs[i].columns = v_arr_jobs[i].iloc[0]
            v_arr_jobs[i] = v_arr_jobs[i].drop([v_arr_jobs[i].index[0]])

        df_job_others = pd.concat(
            [
                v_arr_jobs[1],
                v_arr_jobs[2],
                v_arr_jobs[3],
                v_arr_jobs[4],
                v_arr_jobs[5],
                v_arr_jobs[6],
            ],
            axis=1,
        )

        # insert 날짜
        v_arr_jobs[0].insert(0, "날짜", v_str_cur_date)
        df_job_others.insert(0, "날짜", v_str_cur_date)
        v_arr_jobs[0]["날짜"] = v_arr_jobs[0]["날짜"].apply(
            lambda x: pd.Timestamp(x).strftime("%Y-%m-%d")
        )
        df_job_others["날짜"] = df_job_others["날짜"].apply(
            lambda x: pd.Timestamp(x).strftime("%Y-%m-%d")
        )

        if debug or D_make_df_plan_jobs:
            SUCCESS = f"make_df_plan_jobs : SUCCESS"
            print(f"CODE : {SUCCESS} ")
            print(f"CODE : {v_arr_jobs[0]} ")
            print(f"CODE : {df_job_others} ")

        return v_arr_jobs[0], df_job_others

    except:
        if debug or D_make_df_plan_jobs:
            ERROR = f"make_df_plan_jobs : ERROR 데이터 프레임 만들기 실패"
            print(f"CODE : {ERROR} ")


# FUNCTION CODE : D_make_df_from_arr
#
#
# comment : df 배열을 concat 해서 돌려주는 함수, update 가 아닌 그대로 저장
###############################################################################################
def make_df_from_arr(arr_df_frame):
    v_arr_jobs = arr_df_frame
    v_df_jobs = [[], [], []]

    try:
        for i in range(0, len(v_arr_jobs)):
            v_df_jobs[i] = v_arr_jobs[i][0]
            for j in range(0, len(v_arr_jobs[i])):
                v_df_jobs[i] = pd.concat([v_df_jobs[i], v_arr_jobs[i][j]], axis=0)
                v_df_jobs[i] = v_df_jobs[i].drop_duplicates("날짜").reset_index(drop=True)

        if debug or D_make_df_from_arr:
            SUCCESS = f"make_df_from_arr : SUCCESS"
            print(f"CODE : {SUCCESS} ")
            print(f"CODE : {v_df_jobs} ")

        return v_df_jobs

    except:
        if debug or D_make_df_from_arr:
            ERROR = f"make_df_from_arr : ERROR DateFrame from the Sequence"
            print(f"CODE : {ERROR} ")


# FUNCTION CODE : D_update_df_from_arr
#
#
# comment : df 배열을 concat 해서 돌려주는 함수
###############################################################################################
def update_df_from_arr(arr_df_frame):
    v_arr_jobs = arr_df_frame
    v_df_jobs = [[], [], []]
    v_df_time_jobs = load_time_excel()

    try:
        for i in range(0, len(v_arr_jobs)):
            v_df_jobs[i] = v_df_time_jobs[i]
            for j in range(0, len(v_arr_jobs[i])):
                v_df_jobs[i] = pd.concat([v_df_jobs[i], v_arr_jobs[i][j]], axis=0)
                v_df_jobs[i] = v_df_jobs[i].drop_duplicates("날짜").reset_index(drop=True)
                v_df_jobs[i] = v_df_jobs[i].sort_values("날짜", ascending=False)

        if debug or D_update_df_from_arr:
            SUCCESS = f"update_df_from_arr : SUCCESS"
            print(f"CODE : {SUCCESS} ")
            print(f"CODE : {v_df_jobs[0]} ")

        return v_df_jobs

    except:
        if debug or D_update_df_from_arr:
            ERROR = f"update_df_from_arr : ERROR DateFrame from the Sequence"
            print(f"CODE : {ERROR} ")


# FUNCTION CODE : save_time_excel
# path
# file        :
# comment :
###############################################################################################


def save_time_excel(df1, df2, df3):
    src = r"W:\01. 업무\000. 계획\01. 시간분석테이블/"
    file = "03. 시간분석테이블.xlsx"
    file_b = "03. 시간분석테이블-backup.xlsx"
    src_file = r"W:\01. 업무\000. 계획\01. 시간분석테이블/03. 시간분석테이블.xlsx"

    sht = []
    sht = read_sheets(src, file)

    try:
        if debug or D_save_time_excel:
            SUCCESS = f"save_time_excel : SUCCESS"
            print(f"CODE : {SUCCESS} ")
            print(f"CODE : {sht} ")

        with pd.ExcelWriter(src_file) as writer:
            df1.to_excel(writer, sheet_name=sht[0], index=False)
            df2.to_excel(writer, sheet_name=sht[1], index=False)
            df3.to_excel(writer, sheet_name=sht[2], index=False)

        with pd.ExcelWriter(src + file_b) as writer:
            df1.to_excel(writer, sheet_name=sht[0], index=False)
            df2.to_excel(writer, sheet_name=sht[1], index=False)
            df3.to_excel(writer, sheet_name=sht[2], index=False)

    except:
        if debug or D_save_time_excel:
            ERROR = f"save_time_excel : ERROR 타임시트 엑셀에 저장실패"
            print(f"CODE : {ERROR} ")


try:

    # FUNCTION CODE : make_df_plan_sum
    # path = 파일이 있는 경로명
    # file = 파일명.xlsx
    # src  = path + file
    #
    # comment :
    ###############################################################################################
    def make_df_plan(cmd, cnt_sht_num=1, date="230101"):
        v_str_command = cmd
        v_int_cnt_sht_num = cnt_sht_num
        v_str_sht_date = date

        path = r"W:\01. 업무\000. 계획/"
        file = "01. 2023 일일계획표.xlsx"
        src = path + file
        frame = []
        df_sum_frame = []
        df_job1_frame = []
        df_others_frame = []

        try:
            # get 계획표 from 일일계획표, 현재날짜 + 갯수
            if v_str_command == "load_shts_by_number":
                frame = load_cur_plan_excel(
                    "load_shts_by_number", v_int_cnt_sht_num, v_str_sht_date
                )
            else:
                frame = load_cur_plan_excel(
                    "load_sht_by_date", v_int_cnt_sht_num, v_str_sht_date
                )

            for i in range(0, len(frame)):
                df_plan_sum = make_df_plan_sum(frame[i])
                df_plan_job1, df_plan_others = make_df_plan_jobs(frame[i])
                df_sum_frame.append(df_plan_sum)
                df_job1_frame.append(df_plan_job1)
                df_others_frame.append(df_plan_others)

            if debug or D_make_df_plan:
                SUCCESS = f"make_df_plan : SUCCESS"
                print(f"CODE : {SUCCESS} ")
                print(f"LOAD BY: {v_str_command} ")
                print(f"\tREAD SUM FRAME : {df_sum_frame}")
                print(f"\tREAD SUM FRAME : {df_job1_frame}")
                print(f"\tREAD SUM FRAME : {df_others_frame}")

            return df_sum_frame, df_job1_frame, df_others_frame

        except:
            if debug or D_make_df_plan:
                ERROR = f"make_df_plan : ERROR 없는 시트에 접근 하였음"
                print(f"CODE : {ERROR} ")

    #######################################################################################################
    # MAIN CODE START

    src_path = r"D:\000. 발주서\02. SRC_M/"
    del_sht_name = "Order"
    del_sht_number = 1
    del_sht_start_number = 1
    del_sht_end_number = 2
    SHT_NAME_TYPE = 0
    SHT_NUM_TYPE = 1
    SHT_RNG_TYPE = 2
    SHT_READ_TYPE = 3
    READ_DFE = 1
    v_int_read_sht_qty = 11

    # md.create_new_excel()
    # md.load_time_excel()
    # md.load_cur_plan_excel()
    # md.make_df_plan()
    # make_df_plan(v_int_read_sht_qty)

except Exception as ex:
    print("error" + str(ex))
